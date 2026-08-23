"""
Analyzes the Monthly_Expenditure_Analysis_2021_2026.xlsx workbook (or a
Google-Sheets-exported equivalent) and produces a comprehensive JSON summary.

Usage:
    python execution/analyze_expenditure.py              # print JSON to stdout
    python execution/analyze_expenditure.py --output     # write data/finance_data.json
"""

import argparse
import json
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl

# ── Paths ──────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).parent.parent
WORKBOOK_PATH = BASE_DIR / 'Monthly_Expenditure_Analysis_2021_2026.xlsx'
OUTPUT_PATH   = BASE_DIR / 'data' / 'finance_data.json'

# ── Load workbook ──────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(WORKBOOK_PATH, data_only=True)
ws = wb.active

headers = [cell.value for cell in ws[1]]
rows = []
for row in ws.iter_rows(min_row=2, values_only=True):
    if row[0]:
        rows.append(dict(zip(headers, row)))

# ── Overall totals ──────────────────────────────────────────────────────────────
total_months   = len(rows)
total_income   = sum(r['Total_Income_BDT']   or 0 for r in rows)
total_spent    = sum(r['Total_Spent_BDT']    or 0 for r in rows)
total_salary   = sum(r['Salary_BDT']         or 0 for r in rows)
total_savings  = sum(r['Net_Savings_BDT']    or 0 for r in rows)
total_tx       = int(sum(r['Tx_Count']       or 0 for r in rows))
latest_balance = rows[-1]['Ending_Balance_BDT']
latest_month   = rows[-1]['Year_Month'].strftime('%B %Y')

avg_monthly_spend   = total_spent   / total_months
avg_monthly_income  = total_income  / total_months
avg_monthly_savings = total_savings / total_months

positive_months  = sum(1 for r in rows if (r['Net_Savings_BDT'] or 0) > 0)
negative_months  = total_months - positive_months
savings_rate_pct = (total_savings / total_income * 100) if total_income else 0

# ── Best / worst months ─────────────────────────────────────────────────────────
best_month           = max(rows, key=lambda x: x['Net_Savings_BDT']  or 0)
worst_month          = min(rows, key=lambda x: x['Net_Savings_BDT']  or 0)
highest_spend_month  = max(rows, key=lambda x: x['Total_Spent_BDT']  or 0)
highest_income_month = max(rows, key=lambda x: x['Total_Income_BDT'] or 0)
busiest_month        = max(rows, key=lambda x: x['Tx_Count']         or 0)

# ── Year-by-year breakdown ──────────────────────────────────────────────────────
by_year = defaultdict(list)
for r in rows:
    by_year[r['Year_Month'].year].append(r)

yearly = {}
for yr, rlist in sorted(by_year.items()):
    yearly[str(yr)] = {
        'months':  len(rlist),
        'income':  round(sum(r['Total_Income_BDT'] or 0 for r in rlist), 2),
        'spent':   round(sum(r['Total_Spent_BDT']  or 0 for r in rlist), 2),
        'savings': round(sum(r['Net_Savings_BDT']  or 0 for r in rlist), 2),
        'tx':      int(sum(r['Tx_Count'] or 0 for r in rlist)),
    }

# ── Monthly time-series ─────────────────────────────────────────────────────────
monthly_series = []
for r in rows:
    monthly_series.append({
        'label':             r['Year_Month'].strftime('%b %Y'),
        'iso':               r['Year_Month'].strftime('%Y-%m'),
        'spent':             round(r['Total_Spent_BDT']                                   or 0, 2),
        'income':            round(r['Total_Income_BDT']                                  or 0, 2),
        'salary':            round(r['Salary_BDT']                                        or 0, 2),
        'accrualSalary':     round(r.get('Accrual_Salary_BDT') or r['Salary_BDT']         or 0, 2),
        'savings':           round(r['Net_Savings_BDT']                                   or 0, 2),
        'normalizedSavings': round(r.get('Normalized_Savings_BDT') or r['Net_Savings_BDT'] or 0, 2),
        'balance':           round(r['Ending_Balance_BDT']                                or 0, 2),
        'tx':                int(r['Tx_Count'] or 0),
    })

# ── Assemble summary ────────────────────────────────────────────────────────────
summary = {
    'overview': {
        'total_months':        total_months,
        'total_income':        round(total_income,  2),
        'total_spent':         round(total_spent,   2),
        'total_salary':        round(total_salary,  2),
        'total_savings':       round(total_savings, 2),
        'total_transactions':  total_tx,
        'latest_balance':      round(latest_balance, 2),
        'latest_month':        latest_month,
        'avg_monthly_spend':   round(avg_monthly_spend,   2),
        'avg_monthly_income':  round(avg_monthly_income,  2),
        'avg_monthly_savings': round(avg_monthly_savings, 2),
        'positive_months':     positive_months,
        'negative_months':     negative_months,
        'savings_rate_pct':    round(savings_rate_pct, 2),
    },
    'highlights': {
        'best_month':            best_month['Year_Month'].strftime('%B %Y'),
        'best_month_savings':    round(best_month['Net_Savings_BDT'] or 0, 2),
        'worst_month':           worst_month['Year_Month'].strftime('%B %Y'),
        'worst_month_savings':   round(worst_month['Net_Savings_BDT'] or 0, 2),
        'highest_spend_month':   highest_spend_month['Year_Month'].strftime('%B %Y'),
        'highest_spend_amount':  round(highest_spend_month['Total_Spent_BDT'] or 0, 2),
        'highest_income_month':  highest_income_month['Year_Month'].strftime('%B %Y'),
        'highest_income_amount': round(highest_income_month['Total_Income_BDT'] or 0, 2),
        'busiest_month':         busiest_month['Year_Month'].strftime('%B %Y'),
        'busiest_tx_count':      int(busiest_month['Tx_Count'] or 0),
    },
    'yearly':         yearly,
    'monthly_series': monthly_series,
}

json_str = json.dumps(summary, indent=2)

# ── Output ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser(description='Analyze expenditure data')
parser.add_argument('--output', action='store_true',
                    help='Write JSON to data/finance_data.json instead of stdout')
args = parser.parse_args()

if args.output:
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json_str, encoding='utf-8')
    print(f'Written to {OUTPUT_PATH}  ({len(json_str)} bytes)', file=sys.stderr)
else:
    print(json_str)
